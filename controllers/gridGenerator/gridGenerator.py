# This file is responsible for generating time grids for a new hires class schedule
# The programmatic equivalent to manually highlighting a student's grid on paper
# Generates a new timetable in the project directory

import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import time

# Add the external directory to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
external_directory = os.path.join(current_dir, "..")
sys.path.append(external_directory)

# Import functions from custom modules
from api_calls.workday_api.workday_api import getStudentSchedule
from utils.helperFunctions import convert_to_time

# Get the Excel file path from command-line arguments
if len(sys.argv) < 2:
    print("Error: The path to the Excel file must be provided as an argument.")
    sys.exit(1)

excel_file_path = sys.argv[1]
print(f"Excel file path: {excel_file_path}")

# Check if the file exists
if not os.path.exists(excel_file_path):
    print("Error: The file Timetable.xlsx does not exist at the specified path.")
    sys.exit(1)

# Load the workbook and access the active sheet
wb = load_workbook(excel_file_path)
ws = wb.active

def populateGrid(data):
    """
    Iterate through each day to fill in the grid based on the provided data.
    """
    for course in data:
        fillInDay(course, "U", 3)  # Sunday
        fillInDay(course, "M", 4)  # Monday
        fillInDay(course, "T", 5)  # Tuesday
        fillInDay(course, "W", 6)  # Wednesday
        fillInDay(course, "R", 7)  # Thursday
        fillInDay(course, "F", 8)  # Friday
        fillInDay(course, "S", 9)  # Saturday

def fillInDay(course, day, rowNum):
    """
    Iterates through each time slot in the "day" row of the grid and determines whether or not to highlight that slot.
    """
    max_col = ws.max_column
    hour = 6  # Start at 6 AM
    for col_num_outer in range(2, max_col + 1, 12):  # Loop through each hour
        minute = 0
        for col_num_inner in range(col_num_outer, col_num_outer + 12):  # Loop through each 5-minute increment
            currentTime = time(hour, minute)
            if not isAvailable(course, day, currentTime):
                cell = ws.cell(row=rowNum, column=col_num_inner)
                fillColor = PatternFill(start_color="98FF98", end_color="98FF98", fill_type="solid")  # Light green color
                cell.fill = fillColor
            minute += 5
        hour += 1

def isAvailable(course, day, currentTime):
    """
    Observes one time slot on the Excel sheet and determines whether or not the course meets during that time.
    """
    days = course["meetingDays"]
    startClass = convert_to_time(course["start"])
    endClass = convert_to_time(course["end"])
    for char in days:
        if char == day:
            if currentTime >= startClass and currentTime < endClass:
                return False
    return True

# Comment out or remove the clearGrid function call
# clearGrid()

# Fetch the student schedule data from Workday
data1 = getStudentSchedule(0)

if data1:
    populateGrid(data1)  # Populate the grid with the fetched data
else:
    print("Error fetching data from Workday")

# Save the workbook to the specified path
wb.save(excel_file_path)
